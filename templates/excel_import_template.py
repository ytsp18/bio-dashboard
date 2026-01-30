"""
Excel Import Template Generator for Bio Dashboard
สร้าง Excel Template สำหรับนำเข้าข้อมูลเข้าระบบ Bio Dashboard

Usage:
    python excel_import_template.py [output_path]

Example:
    python excel_import_template.py "Bio_import_template.xlsx"
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
import os
import sys


# ==================== SHEET DEFINITIONS ====================

SHEETS = {
    # Sheet 1: สรุปภาพรวม (Summary) - สำหรับอ้างอิง ไม่ต้อง import
    '1.สรุปภาพรวม': {
        'description': 'สรุปผลภาพรวม (ไม่ต้องกรอก - สร้างอัตโนมัติ)',
        'columns': [],
        'skip_import': True,
    },

    # Sheet 2: รายการบัตรดี
    '2.รายการบัตรดี': {
        'description': 'รายละเอียดบัตรดี (G)',
        'columns': [
            {'name': 'ลำดับ', 'type': 'int', 'required': False, 'width': 8},
            {'name': 'Appointment ID', 'type': 'text', 'required': True, 'width': 22, 'example': '1-TRT001012600040'},
            {'name': 'รหัสศูนย์', 'type': 'text', 'required': True, 'width': 15, 'example': 'TRT-SC-S-001'},
            {'name': 'ชื่อศูนย์', 'type': 'text', 'required': False, 'width': 45, 'example': 'ศูนย์บริการใบอนุญาตทำงานต่างจังหวัด จังหวัดตราด'},
            {'name': 'ภูมิภาค', 'type': 'text', 'required': False, 'width': 15, 'example': 'ภาคตะวันออก'},
            {'name': 'Card ID', 'type': 'text_13', 'required': True, 'width': 18, 'example': '6923000000006'},
            {'name': 'Serial Number', 'type': 'text_13', 'required': True, 'width': 18, 'example': '0099004524052'},
            {'name': 'Work Permit No', 'type': 'text_13', 'required': False, 'width': 18, 'example': '0769230000001'},
            {'name': 'SLA (นาที)', 'type': 'float', 'required': False, 'width': 12, 'example': '5.5'},
            {'name': 'ผ่าน SLA', 'type': 'text', 'required': False, 'width': 10, 'example': 'ผ่าน'},
            {'name': 'ผู้ให้บริการ', 'type': 'text', 'required': False, 'width': 20, 'example': 'apisara.yus'},
            {'name': 'วันที่พิมพ์', 'type': 'date', 'required': True, 'width': 15, 'example': '2026-01-06'},
        ],
    },

    # Sheet 3: รายการบัตรเสีย
    '3.รายการบัตรเสีย': {
        'description': 'รายละเอียดบัตรเสีย (B)',
        'columns': [
            {'name': 'ลำดับ', 'type': 'int', 'required': False, 'width': 8},
            {'name': 'Appointment ID', 'type': 'text', 'required': True, 'width': 22, 'example': '1-TRT001012600040'},
            {'name': 'รหัสศูนย์', 'type': 'text', 'required': True, 'width': 15, 'example': 'TRT-SC-S-001'},
            {'name': 'ชื่อศูนย์', 'type': 'text', 'required': False, 'width': 45},
            {'name': 'ภูมิภาค', 'type': 'text', 'required': False, 'width': 15},
            {'name': 'Card ID', 'type': 'text_13', 'required': True, 'width': 18},
            {'name': 'Serial Number', 'type': 'text_13', 'required': True, 'width': 18},
            {'name': 'สาเหตุ', 'type': 'text', 'required': False, 'width': 30, 'example': 'พิมพ์ผิด'},
            {'name': 'ผู้ให้บริการ', 'type': 'text', 'required': False, 'width': 20},
            {'name': 'วันที่พิมพ์', 'type': 'date', 'required': True, 'width': 15},
        ],
    },

    # Sheet 4: สรุปตามศูนย์
    '4.สรุปตามศูนย์': {
        'description': 'สถิติแยกตามศูนย์บริการ',
        'columns': [
            {'name': 'ลำดับ', 'type': 'int', 'required': False, 'width': 8},
            {'name': 'รหัสศูนย์', 'type': 'text', 'required': True, 'width': 15},
            {'name': 'ชื่อศูนย์', 'type': 'text', 'required': False, 'width': 45},
            {'name': 'จำนวนบัตรดี', 'type': 'int', 'required': True, 'width': 15},
            {'name': 'SLA เฉลี่ย', 'type': 'float', 'required': False, 'width': 12},
            {'name': 'SLA สูงสุด', 'type': 'float', 'required': False, 'width': 12},
        ],
    },

    # Sheet 7: บัตรจัดส่ง
    '7.บัตรจัดส่ง': {
        'description': 'รายการบัตรจัดส่ง',
        'columns': [
            {'name': 'ลำดับ', 'type': 'int', 'required': False, 'width': 8},
            {'name': 'Appointment ID', 'type': 'text', 'required': True, 'width': 22},
            {'name': 'Serial Number', 'type': 'text_13', 'required': True, 'width': 18},
            {'name': 'สถานะ', 'type': 'text', 'required': True, 'width': 10, 'example': 'G'},
            {'name': 'Card ID', 'type': 'text_13', 'required': False, 'width': 18},
            {'name': 'Work Permit No', 'type': 'text_13', 'required': False, 'width': 18},
        ],
    },

    # Sheet 9: ออกบัตรผิดศูนย์
    '9.ออกบัตรผิดศูนย์': {
        'description': 'รายที่ออกบัตรผิดศูนย์นัด',
        'columns': [
            {'name': 'ลำดับ', 'type': 'int', 'required': False, 'width': 8},
            {'name': 'Appointment ID', 'type': 'text', 'required': True, 'width': 22},
            {'name': 'ศูนย์ที่นัด', 'type': 'text', 'required': True, 'width': 15},
            {'name': 'ศูนย์ที่ออกบัตร', 'type': 'text', 'required': True, 'width': 15},
            {'name': 'Serial Number', 'type': 'text_13', 'required': False, 'width': 18},
            {'name': 'สถานะ', 'type': 'text', 'required': False, 'width': 10},
            {'name': 'วันที่พิมพ์', 'type': 'date', 'required': False, 'width': 15},
        ],
    },

    # Sheet 13: ข้อมูลทั้งหมด (Full data)
    '13.ข้อมูลทั้งหมด': {
        'description': 'ข้อมูล Raw ทั้งหมด',
        'columns': [
            {'name': 'ลำดับ', 'type': 'int', 'required': False, 'width': 8},
            {'name': 'Appointment ID', 'type': 'text', 'required': True, 'width': 22},
            {'name': 'Form ID', 'type': 'text', 'required': False, 'width': 15},
            {'name': 'Form Type', 'type': 'text', 'required': False, 'width': 12},
            {'name': 'Branch Code', 'type': 'text', 'required': True, 'width': 15},
            {'name': 'Card ID', 'type': 'text_13', 'required': True, 'width': 18},
            {'name': 'Work Permit No', 'type': 'text_13', 'required': False, 'width': 18},
            {'name': 'Serial Number', 'type': 'text_13', 'required': True, 'width': 18},
            {'name': 'Print Status', 'type': 'text', 'required': True, 'width': 12, 'example': 'G'},
            {'name': 'Reject Type', 'type': 'text', 'required': False, 'width': 20},
            {'name': 'Operator', 'type': 'text', 'required': False, 'width': 20},
            {'name': 'Print Date', 'type': 'date', 'required': True, 'width': 15},
            {'name': 'SLA Start', 'type': 'datetime', 'required': False, 'width': 20},
            {'name': 'SLA Stop', 'type': 'datetime', 'required': False, 'width': 20},
            {'name': 'SLA Duration', 'type': 'text', 'required': False, 'width': 12},
            {'name': 'SLA Minutes', 'type': 'float', 'required': False, 'width': 12},
            {'name': 'Branch Name', 'type': 'text', 'required': False, 'width': 45},
            {'name': 'Region', 'type': 'text', 'required': False, 'width': 15},
        ],
    },

    # Sheet 22: ส่วนต่างบัตรสมบูรณ์
    '22.ส่วนต่างบัตรสมบูรณ์': {
        'description': 'ส่วนต่างระหว่าง Unique SN (G) กับบัตรสมบูรณ์',
        'columns': [
            {'name': 'ลำดับ', 'type': 'int', 'required': False, 'width': 8},
            {'name': 'Appointment ID', 'type': 'text', 'required': True, 'width': 22},
            {'name': 'จำนวน G', 'type': 'int', 'required': False, 'width': 10},
            {'name': 'รหัสศูนย์', 'type': 'text', 'required': True, 'width': 15},
            {'name': 'ชื่อศูนย์', 'type': 'text', 'required': False, 'width': 45},
            {'name': 'ภูมิภาค', 'type': 'text', 'required': False, 'width': 15},
            {'name': 'Card ID', 'type': 'text_13', 'required': False, 'width': 18},
            {'name': 'Serial Number', 'type': 'text_13', 'required': False, 'width': 18},
            {'name': 'Work Permit No', 'type': 'text_13', 'required': False, 'width': 18},
            {'name': 'SLA (นาที)', 'type': 'float', 'required': False, 'width': 12},
            {'name': 'ผู้ให้บริการ', 'type': 'text', 'required': False, 'width': 20},
            {'name': 'วันที่พิมพ์', 'type': 'date', 'required': False, 'width': 15},
        ],
    },
}


# ==================== STYLES ====================

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
REQUIRED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
EXAMPLE_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_template(output_path: str = 'Bio_import_template.xlsx'):
    """สร้าง Excel Template สำหรับนำเข้าข้อมูล"""

    wb = Workbook()

    # ลบ sheet เริ่มต้น
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # สร้าง sheet คำแนะนำ
    create_instructions_sheet(wb)

    # สร้าง sheets ตาม definition
    for sheet_name, sheet_def in SHEETS.items():
        if sheet_def.get('skip_import'):
            continue

        ws = wb.create_sheet(title=sheet_name)
        create_data_sheet(ws, sheet_def)

    # Save
    wb.save(output_path)
    print(f"✅ สร้าง Template สำเร็จ: {output_path}")
    return output_path


def create_instructions_sheet(wb: Workbook):
    """สร้าง sheet คำแนะนำการใช้งาน"""

    ws = wb.create_sheet(title='คำแนะนำ', index=0)

    instructions = [
        ['Bio Dashboard - Excel Import Template'],
        [''],
        ['คำแนะนำการใช้งาน:'],
        ['1. กรอกข้อมูลในแต่ละ Sheet ตาม format ที่กำหนด'],
        ['2. Column ที่มีพื้นหลังสีแดงอ่อน (*) คือ required fields'],
        ['3. แถวที่ 2 (สีเขียว) คือตัวอย่างข้อมูล - ให้ลบก่อน import'],
        ['4. ตัวเลข 13 หลัก (Card ID, Serial Number, Work Permit No) ต้องใส่เป็น Text'],
        ['5. วันที่ใช้ format: YYYY-MM-DD (เช่น 2026-01-06)'],
        [''],
        ['Sheets ที่รองรับ:'],
        ['- 2.รายการบัตรดี: รายละเอียดบัตรดี (G)'],
        ['- 3.รายการบัตรเสีย: รายละเอียดบัตรเสีย (B)'],
        ['- 4.สรุปตามศูนย์: สถิติแยกตามศูนย์บริการ'],
        ['- 7.บัตรจัดส่ง: รายการบัตรจัดส่ง'],
        ['- 9.ออกบัตรผิดศูนย์: รายที่ออกบัตรผิดศูนย์นัด'],
        ['- 13.ข้อมูลทั้งหมด: ข้อมูล Raw ทั้งหมด'],
        ['- 22.ส่วนต่างบัตรสมบูรณ์: ส่วนต่างระหว่าง Unique SN (G) กับบัตรสมบูรณ์'],
        [''],
        ['หมายเหตุ:'],
        ['- Sheet "1.สรุปภาพรวม" จะถูกสร้างอัตโนมัติจากข้อมูล'],
        ['- ถ้ามีปัญหา column ไม่ตรง ให้ตรวจสอบ header row'],
        [''],
        ['สร้างโดย: Bio Dashboard Import Template Generator'],
        [f'วันที่สร้าง: {date.today().strftime("%Y-%m-%d")}'],
    ]

    for row_idx, row_data in enumerate(instructions, start=1):
        cell = ws.cell(row=row_idx, column=1, value=row_data[0] if row_data else '')
        if row_idx == 1:
            cell.font = Font(bold=True, size=16)
        elif row_data and row_data[0].startswith('-'):
            cell.font = Font(size=10)

    ws.column_dimensions['A'].width = 80


def create_data_sheet(ws, sheet_def: dict):
    """สร้าง data sheet พร้อม header และตัวอย่าง"""

    columns = sheet_def.get('columns', [])
    if not columns:
        return

    # Header row
    for col_idx, col_def in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_def['name'])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

        # Set column width
        ws.column_dimensions[cell.column_letter].width = col_def.get('width', 15)

        # Mark required columns
        if col_def.get('required'):
            # Add asterisk to header
            cell.value = f"{col_def['name']} *"

    # Example row (row 2)
    for col_idx, col_def in enumerate(columns, start=1):
        example = col_def.get('example', '')
        cell = ws.cell(row=2, column=col_idx, value=example)
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='left')

        # Format text_13 columns as text
        if col_def.get('type') == 'text_13':
            cell.number_format = '@'

    # Add data validation for Print Status
    for col_idx, col_def in enumerate(columns, start=1):
        if col_def['name'] in ['Print Status', 'สถานะ']:
            dv = DataValidation(type='list', formula1='"G,B"', allow_blank=True)
            dv.error = 'กรุณาเลือก G หรือ B'
            dv.errorTitle = 'Invalid Status'
            ws.add_data_validation(dv)
            dv.add(ws.cell(row=2, column=col_idx))

    # Freeze header row
    ws.freeze_panes = 'A2'


def create_simple_template(output_path: str = 'Bio_simple_import.xlsx'):
    """
    สร้าง Simple Template - มีเฉพาะ Sheet หลักที่จำเป็น
    เหมาะสำหรับการ import ข้อมูลแบบง่าย
    """

    wb = Workbook()

    # ลบ sheet เริ่มต้น
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Sheet 1: ข้อมูลบัตรทั้งหมด (รวม G และ B)
    ws = wb.create_sheet(title='ข้อมูลบัตร')

    columns = [
        {'name': 'Appointment ID *', 'width': 22, 'example': '1-TRT001012600040'},
        {'name': 'Branch Code *', 'width': 15, 'example': 'TRT-SC-S-001'},
        {'name': 'Branch Name', 'width': 45, 'example': 'ศูนย์บริการฯ จ.ตราด'},
        {'name': 'Card ID *', 'width': 18, 'example': '6923000000006'},
        {'name': 'Serial Number *', 'width': 18, 'example': '0099004524052'},
        {'name': 'Work Permit No', 'width': 18, 'example': '0769230000001'},
        {'name': 'Print Status *', 'width': 12, 'example': 'G'},
        {'name': 'Reject Reason', 'width': 25, 'example': ''},
        {'name': 'Operator', 'width': 20, 'example': 'apisara.yus'},
        {'name': 'SLA Minutes', 'width': 12, 'example': '5.5'},
        {'name': 'Print Date *', 'width': 15, 'example': '2026-01-06'},
        {'name': 'Region', 'width': 15, 'example': 'ภาคตะวันออก'},
    ]

    # Header
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col['name'])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = col['width']

    # Example row
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col.get('example', ''))
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER
        # Format ID columns as text
        if 'ID' in col['name'] or 'Number' in col['name'] or 'Permit' in col['name']:
            cell.number_format = '@'

    # Print Status validation
    dv = DataValidation(type='list', formula1='"G,B"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f'G2:G1000')

    ws.freeze_panes = 'A2'

    # Sheet 2: คำแนะนำ
    ws_help = wb.create_sheet(title='คำแนะนำ')
    help_text = [
        ['Bio Dashboard - Simple Import Template'],
        [''],
        ['Column ที่มี * คือ required fields'],
        [''],
        ['Column Definitions:'],
        ['- Appointment ID: รหัสนัดหมาย (เช่น 1-TRT001012600040)'],
        ['- Branch Code: รหัสศูนย์ (เช่น TRT-SC-S-001)'],
        ['- Branch Name: ชื่อศูนย์ (optional)'],
        ['- Card ID: เลขทะเบียนบัตร 13 หลัก'],
        ['- Serial Number: หมายเลขบัตร 13 หลัก'],
        ['- Work Permit No: เลขใบอนุญาตทำงาน 13 หลัก'],
        ['- Print Status: สถานะ G=บัตรดี, B=บัตรเสีย'],
        ['- Reject Reason: สาเหตุบัตรเสีย (เฉพาะ B)'],
        ['- Operator: ชื่อผู้ให้บริการ'],
        ['- SLA Minutes: เวลาออกบัตร (นาที)'],
        ['- Print Date: วันที่พิมพ์ (YYYY-MM-DD)'],
        ['- Region: ภูมิภาค'],
        [''],
        ['หมายเหตุ:'],
        ['- ลบแถวตัวอย่าง (แถว 2 สีเขียว) ก่อน import'],
        ['- ตัวเลข 13 หลักต้องใส่เป็น Text (เริ่มด้วย \')'],
    ]

    for row_idx, row_data in enumerate(help_text, start=1):
        ws_help.cell(row=row_idx, column=1, value=row_data[0] if row_data else '')

    ws_help.column_dimensions['A'].width = 60

    wb.save(output_path)
    print(f"✅ สร้าง Simple Template สำเร็จ: {output_path}")
    return output_path


if __name__ == '__main__':
    output_dir = os.path.dirname(os.path.abspath(__file__))

    # สร้าง Full Template
    full_template = os.path.join(output_dir, 'Bio_import_template.xlsx')
    create_template(full_template)

    # สร้าง Simple Template
    simple_template = os.path.join(output_dir, 'Bio_simple_import.xlsx')
    create_simple_template(simple_template)

    print("\n📁 Templates created:")
    print(f"   1. {full_template}")
    print(f"   2. {simple_template}")
